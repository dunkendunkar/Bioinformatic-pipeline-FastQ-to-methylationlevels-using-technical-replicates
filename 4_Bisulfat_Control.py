import subprocess
import os
import sys
import glob
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from hypothesis.internal.conjecture.junkdrawer import endswith
from matplotlib.gridspec import GridSpec
import tkinter as tk
from tkinter import filedialog as fd, filedialog
import openpyxl
from matplotlib.ticker import ScalarFormatter
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

print("Bisulfite Control!")
root = tk.Tk()
root.withdraw()
from pathlib import Path
BASE = Path(__file__).resolve().parent
# =======================================================================================================================
#BISFULITE CONVERSION >99% control. Choose two unmethylated, C and make choose BED-file.
# =======================================================================================================================
folder = BASE/"RESULTS"
#files_BAM = filedialog.askdirectory(title= "select folder which has the generated BAM-files")
files_BAM = BASE/"BAM"
#reference_file = filedialog.askopenfilename(title = "select folder with models reference, FASTA-file!")
reference_file = sys.argv[1]
bed_file = filedialog.askopenfilename(title= "select bed-file for Bisulfate control positions")
# Get all BAM files in that folder
bam_files = glob.glob(f"{files_BAM}/*.bam")
out_readcount = BASE/"RESULTS"
for bam_file_sorted in bam_files:
    p = Path(bam_file_sorted)
    base = p.stem              # removes .bam
    base = Path(base).stem     # would remove second extension if present
    sample_name = base
    bai_file = f"{files_BAM}/{sample_name}.bam.bai"
    output_file = f"{out_readcount}/readcount_{sample_name}.txt"
    cmd_bamreadcount = [
        "bam-readcount",
        "--max-warnings", "0",
        "--min-base-quality", "30",
        "--min-mapping-quality", "30",
        "--site-list", bed_file,
        "-f", reference_file,
        bam_file_sorted
    ]
    with open(output_file, "w") as fh:
        subprocess.run(cmd_bamreadcount, stdout=fh, check=True)
    print(f"Saved readcount_{sample_name}, success")
# =======================================================================================================================
#BISFULITE CONVERSION: Extract DATA from readcount files.
# =======================================================================================================================
all_rows = []
files = BASE/"RESULTS"
for file in os.listdir(files):
    if file.startswith("readcount_"):
        path_file = os.path.join(files, file)
        with open(path_file,"r") as fh:
            for line in fh:
                line=line.rstrip("\n")
                if not line:
                    continue
                parts = line.split("\t") #alla positioner som standrad. satt 10 max här. blir annars 200
                gene = parts[0]
                position = int(parts[1])
                ref_base = parts[2]
                total_depth = int(parts[3])
                stats_block = parts[4:]
                row = {"gene": gene,
                       "position": position,
                       "ref_base": ref_base,
                       "total_depth" : total_depth,
                       "source" : file}
                #extract read counts for A C T G N
                for block in stats_block:
                    if block.startswith(("A:","C:","G:","T:","N:")):
                        fields = block.split(":")
                        base = fields[0] #A C G T, definition.
                        read_count = int(fields[1]) #first integer after base
                        row[f"{base}_reads"] = read_count
                all_rows.append(row)
df = pd.DataFrame(all_rows)
print(df.columns.tolist()) # den printar ut 200 coloumns
print(f"exported data in readcounts to dataframe from {folder}")
print(df)
# =======================================================================================================================
#BISFULITE CONVERSION RATE CONTROL:
# =======================================================================================================================
df["%_Bisulfate_conversion"] = 100*(df["T_reads"]/(df["C_reads"]+df["T_reads"]))
df["gene_pos"] = df["gene"].astype(str) + "_" + df["position"].astype(str)
df_control = df[["gene_pos","%_Bisulfate_conversion","total_depth","source"]]
df_control.to_excel(BASE/"RESULTS/bisulfate_control_readcounts.xlsx",index = False)
# =======================================================================================================================
#BISFULITE CONVERSION RATE CONTROL: Color coordinate depending on threshold values in excel file!
# =======================================================================================================================
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

excel_path = BASE/"RESULTS/Bisulfate_controll_readcounts.xlsx"
# Save df_control to Excel
df_control.to_excel(excel_path, index=False)
# Load workbook
wb = load_workbook(excel_path)
ws = wb.active
# Colors
green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
red_fill   = PatternFill(start_color="FFC7CE", fill_type="solid")
# --- 1. %_Bisulfate_conversion column ---
conv_idx = list(df_control.columns).index("%_Bisulfate_conversion") + 1
conv_letter = ws.cell(row=1, column=conv_idx).column_letter
ws.conditional_formatting.add(
    f"{conv_letter}2:{conv_letter}{ws.max_row}",
    CellIsRule(operator="greaterThanOrEqual", formula=["99"], fill=green_fill)
)
ws.conditional_formatting.add(
    f"{conv_letter}2:{conv_letter}{ws.max_row}",
    CellIsRule(operator="lessThan", formula=["99"], fill=red_fill)
)
# --- 2. total_depth column ---
depth_idx = list(df_control.columns).index("total_depth") + 1
depth_letter = ws.cell(row=1, column=depth_idx).column_letter
ws.conditional_formatting.add(
    f"{depth_letter}2:{depth_letter}{ws.max_row}",
    CellIsRule(operator="lessThan", formula=["1000"], fill=red_fill)
)
# Save back to the SAME file
wb.save(excel_path)
print("Completed BISULFITE CONTROL")


